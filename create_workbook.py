from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: Customers ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Customers"
customers_headers = ["customer_id","first_name","last_name","policy_number",
                     "date_of_birth","registered_phone_last4","preferred_language","policy_status"]
customers_rows = [
    ["CUS-1001","Asha","Sharma","POL-1001","1990-02-14","4821","en-US","Active"],
    ["CUS-1002","Rahul","Mehta","POL-1002","1985-07-23","1198","en-US","Active"],
    ["CUS-1003","Neha","Iyer","POL-1003","1992-11-05","7744","en-US","Active"],
    ["CUS-1004","Vikram","Rao","POL-1004","1978-03-18","2056","en-US","Active"],
    ["CUS-1005","Mira","Patel","POL-1005","1965-09-30","6612","en-US","Active"],
    ["CUS-1006","Dev","Kapoor","POL-1006","1988-12-08","9043","en-US","Lapsed"],
]
ws1.append(customers_headers)
for row in customers_rows:
    ws1.append(row)
t1 = Table(displayName="Customers", ref=f"A1:{get_column_letter(len(customers_headers))}{len(customers_rows)+1}")
t1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws1.add_table(t1)

# ── Sheet 2: Claims ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Claims")
claims_headers = ["claim_number","customer_id","policy_number","claim_type","incident_date",
                  "filed_date","status","status_explanation","missing_documents",
                  "next_action","expected_update_date","last_updated","escalation_eligible"]
claims_rows = [
    ["CLM-1042","CUS-1001","POL-1001","Collision","2026-08-03","2026-08-04","Under review",
     "The claims team is reviewing the incident information and vehicle assessment.","",
     "Wait for the assessment review to finish.","2026-08-15","2026-08-12","false"],
    ["CLM-1043","CUS-1002","POL-1002","Collision","2026-08-01","2026-08-02","Information required",
     "The claims team needs an additional document before assessment can continue.","Police incident report",
     "Provide the police incident report through the approved document channel.","2026-08-17","2026-08-12","false"],
    ["CLM-1044","CUS-1003","POL-1003","Weather damage","2026-08-05","2026-08-05","Assessment scheduled",
     "A vehicle assessment has been arranged.","",
     "Attend the assessment appointment listed in the official claim communication.","2026-08-16","2026-08-11","false"],
    ["CLM-1045","CUS-1004","POL-1004","Theft","2026-07-27","2026-07-27","Payment processing",
     "The approved payment is being prepared.","",
     "Wait for the recorded payment processing step to complete.","2026-08-14","2026-08-12","false"],
    ["CLM-1046","CUS-1005","POL-1005","Collision","2026-07-20","2026-07-21","Denied",
     "The claim record shows a decision that requires a claims representative to explain.","",
     "Speak with a claims representative about the recorded decision.","","2026-08-10","true"],
    ["CLM-1047","CUS-1001","POL-1001","Glass damage","2026-06-11","2026-06-11","Closed",
     "The claims team has completed its current activity on this claim.","",
     "Contact a claims representative if you need to discuss reopening or disputing the claim.","","2026-06-20","true"],
]
ws2.append(claims_headers)
for row in claims_rows:
    ws2.append(row)
t2 = Table(displayName="Claims", ref=f"A1:{get_column_letter(len(claims_headers))}{len(claims_rows)+1}")
t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws2.add_table(t2)

# ── Sheet 3: Callbacks ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("Callbacks")
callbacks_headers = ["callback_reference","customer_id","claim_number","callback_reason",
                     "preferred_date","preferred_window","masked_phone","status","created_at","correlation_id"]
callbacks_rows = [
    ["CB-20260814-0999","CUS-1004","CLM-1045","Question about recorded payment timing",
     "2026-08-14","09:00-12:00","******2056","Open","2026-08-13T06:30:00Z",
     "00000000-0000-4000-8000-000000000999"],
]
ws3.append(callbacks_headers)
for row in callbacks_rows:
    ws3.append(row)
t3 = Table(displayName="Callbacks", ref=f"A1:{get_column_letter(len(callbacks_headers))}{len(callbacks_rows)+1}")
t3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws3.add_table(t3)

wb.save("data/Voice360DemoData.xlsx")
print("Created: data/Voice360DemoData.xlsx")
